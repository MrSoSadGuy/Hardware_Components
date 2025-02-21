from flask import json
from flask_login import current_user
from models import *
from pon_models import *
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from getData import *


# def create_file_sostav(name_PON):
#     unit1 = Unit.query.filter_by(name_PON=name_PON).all()
#     if "UL" in name_PON:
#         book = openpyxl.load_workbook('files for download\шаблон.xlsx')
#         sheet = book['Huawei']
#         del book['ZTE']
#         plata_row = 5
#     if "OL" in name_PON:
#         book = openpyxl.load_workbook('files for download\шаблон.xlsx')
#         sheet = book['ZTE']
#         del book['Huawei']
#         plata_row = 6
#     for un in unit1:
#         row = int(un.plata_mesto) + plata_row
#         sheet['B' + str(row)] = un.inv_number
#         sheet['C' + str(row)] = un.name_unit
#         sheet['D' + str(row)] = un.serial_number
#     book.save("files for download\Состав оборудования " + name_PON + ".xlsx")
#     return "files for download\Состав оборудования " + name_PON + ".xlsx"


def create_file_sostav(id):
    unit = List_of_olt.query.get_or_404(int(id))
    if unit.type_of_olt == 1:
        book = openpyxl.load_workbook('files for download\шаблон.xlsx')
        sheet = book['ZTE']
        del book['Huawei']
        plata_row = 6
        
    if unit.type_of_olt == 2 or unit.type_of_olt == 3:
        book = openpyxl.load_workbook('files for download\шаблон.xlsx')
        sheet = book['Huawei']
        del book['ZTE']
        plata_row = 5
    sheet['B4'] = unit.inv_number
    sheet['C4'] = unit.name
    sheet['D4'] = unit.serial_number
    for un in unit.list_of_modules:
        print(un.Olt_sockets.socket)
        row = int(un.Olt_sockets.socket) + plata_row
        sheet['B' + str(row)] = un.inv_number
        sheet['C' + str(row)] = un.name_of_modules
        sheet['D' + str(row)] = un.serial_number
    book.save("files for download\Состав оборудования.xlsx")
    return "files for download\Состав оборудования.xlsx"


def create_file_KTS(id):
    kts = List_of_olt.query.get_or_404(id)
    # kts = Data_for_KTS.query.filter_by(cod_name=name_PON).first()
    user = Users.query.get(current_user.get_id())
    book = openpyxl.load_workbook('files for download\КТС_шаблон.xlsx')
    sheet = book.active
    sheet['A3'] = kts.full_name
    sheet['A4'] = kts.cod_name_of_olt
    sheet['A8'] = kts.Uzel_dostupa.name + ' ' + kts.Uzel_dostupa.Adress + ", " + kts.row_box_shelf + ", ip: " + kts.IP
    sheet['E15'] = kts.zavod
    sheet['L18'] = kts.date_of_production
    sheet['G21'] = kts.serial_number
    sheet['L24'] = kts.inv_number
    sheet['L27'] = kts.date_of_entry
    sheet['L30'] = str(datetime.now().strftime("%d/%m/%Y"))
    sheet['J39'] = user.FIO
    book.save("files for download\КТС.xlsx")
    return "files for download\КТС.xlsx"


def createBuhDataFile(list_data):
    start_row = 4
    row_count = 1
    step = 1
    path = 'files for download\шаблон Бухгалтерские данные.xlsx'
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet = wb_obj.active
        thins = Side(border_style="thin", color="000000")
        double = Side(border_style="medium", color="000000")
        for id in list_data:
            
            buh = BuhUch.query.get_or_404(int(id))
            sheet["A" + str(start_row)] = row_count
            sheet["A" + str(start_row)].border = Border(top=double, bottom=thins, left=double, right=thins)
            sheet["B" + str(start_row)] = buh.inv_number
            sheet["B" + str(start_row)].alignment = Alignment(wrap_text=True)
            sheet["B" + str(start_row)].border = Border(top=double, bottom=thins, left=thins, right=thins)
            sheet["C" + str(start_row)] = buh.name
            sheet["C" + str(start_row)].alignment = Alignment(wrap_text=True)
            sheet["C" + str(start_row)].border = Border(top=double, bottom=thins, left=thins, right=thins)
            sheet["D" + str(start_row)] = buh.MOL
            sheet["D" + str(start_row)].border = Border(top=double, bottom=thins, left=thins, right=double)                       
            if (type(buh.charracter) == str):
                for i in range(0,len(buh.charracter.split('\n'))):
                    start = "A" + str(start_row+1+i)
                    end = "D" + str(start_row+1+i)
                    sheet.merge_cells(start+':'+end)
                    sheet["A" + str(start_row+1+i)] = buh.charracter.split('\n')[i]
                    step = len(buh.charracter.split('\n'))
            else:
                start = "A" + str(start_row+1)
                end = "D" + str(start_row+1)
                sheet.merge_cells(start+':'+end) 
                sheet["A" + str(start_row+1)] = "Нет данных!"
            start_row = start_row + 2 + step
            row_count = row_count + 1
            step = 1
        wb_obj.save('files for download\Бухгалтерские данные.xlsx')
        return json.dumps("SUCCESS")
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")
        return json.dumps(f"Unexpected {err=}, {type(err)=}")
    
def createPONDatafile(list_data):
    start_row = 4
    path = 'files for download\шаблон Таблица оборудования PON.xlsx'
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet = wb_obj.active
        for row in list_data:
            if(row[1] == 'List_of_olt'):
                print(row[0])
                olt = List_of_olt.query.get_or_404(str(row[0]))
                N_ud = olt.Uzel_dostupa.number_ud
                adr_ud = olt.Uzel_dostupa.Adress 
                cod = olt.cod_name_of_olt
                if  'СТЕЛАЖ' not in olt.name:
                    name = olt.name 
                else: continue
                inv_number = olt.inv_number
                ser_num = olt.serial_number
                p_mesto = ''
                note = olt.note
                r_mesto = olt.kts.mesto if olt.kts else ""
            if(row[1] == 'List_of_modules'):
                print(row[0])
                mod = List_of_modules.query.get_or_404(str(row[0]))
                N_ud = mod.List_of_olt.Uzel_dostupa.number_ud
                adr_ud = mod.List_of_olt.Uzel_dostupa.Adress
                cod = mod.List_of_olt.cod_name_of_olt
                name = mod.name_of_modules
                inv_number = mod.inv_number
                ser_num = mod.serial_number
                p_mesto = mod.Olt_sockets.socket if mod.Olt_sockets else ""
                note = mod.note
                r_mesto = mod.List_of_olt.kts.mesto if mod.List_of_olt.kts else ""
            sheet["B" + str(start_row)] = N_ud + ' ' + adr_ud
            sheet["C" + str(start_row)] = cod
            sheet["D" + str(start_row)] = name
            sheet["E" + str(start_row)] = inv_number
            sheet["F" + str(start_row)] = ser_num
            sheet["A" + str(start_row)] = r_mesto
            sheet["G" + str(start_row)] = p_mesto
            sheet["H" + str(start_row)] = note
            start_row = start_row + 1
        wb_obj.save('files for download\Таблица оборудования PON.xlsx')
        return json.dumps("SUCCESS")
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")
        return json.dumps(f"Unexpected {err=}, {type(err)=}")