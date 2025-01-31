from flask import render_template, request, json, send_file, jsonify, redirect, url_for
import openpyxl
from openpyxl.styles import Border, Side, Alignment                                                                  
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from app import app
from models import *
from saveData import *
from delData import *
from pon_models import *
from getData import *

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main'))
    else:
        return redirect(url_for('login'))


@app.route('/login')
def login():
    return render_template("login.html")


@app.route('/logout')
@login_required
def logout():
    user = Users.query.get_or_404(current_user.get_id())
    print(str(datetime.now())+": Пользователь "+user.FIO+" вышел")
    logout_user()
    return redirect(url_for('login'), 301)


@app.route('/login/check_user', methods=['GET', 'POST'])
def login_page():
    req = request.form['json']
    login_data = json.loads(req)
    login = login_data['login']
    password = login_data['password']
    if (login != "") and (password != ""):
        user = Users.query.filter_by(login=login).scalar()
        if user is None:
            return json.dumps('Логин введен неправильно!')
        else:
            if check_password_hash(user.password, password):
                login_user(user)
                print(str(datetime.now())+": Пользователь "+user.FIO+" вошел")
                return json.dumps('SUCCESS')
            else:
                return json.dumps('Пароль введен неправильно!')
    else:
        return json.dumps('Не заполнены  поля логина или пароля')


@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    request_new_pass = request.form['json']
    old_pass = str(json.loads(request_new_pass)["old_pass"])
    new_pass = str(json.loads(request_new_pass)["new_pass"])
    new_pass_2 = str(json.loads(request_new_pass)["new_pass_2"])
    user = Users.query.get_or_404(current_user.get_id())
    print(user.FIO)
    if check_password_hash(user.password, old_pass):
        if new_pass == new_pass_2:
            user.password = generate_password_hash(new_pass)
            return save_data_to_db()
        else:
            return jsonify("Новые пароли не совпадают")
    else:
        return jsonify("Старый пароль введен не верно")


@app.route('/main')
@login_required
def main():
    user = Users.query.get_or_404(current_user.get_id())
    user_name = user.FIO
    return render_template("start_page.html", user_name=user_name)


@app.route('/pon_units')
@login_required
def pon_page():
    units = Unit.query.order_by(Unit.name_PON).all()
    kts_data = Data_for_KTS.query.all()
    user = Users.query.get_or_404(current_user.get_id())
    user_name = user.FIO
    mols = MOLs.query.all()
    return render_template("index.html", units=units, user_name=user_name, kts_data=kts_data, mols = mols)


@app.route('/pon_units_new')
@login_required
def pon_page_new():
    ud = Uzel_dostupa.query.order_by(Uzel_dostupa.id).all()
    user = Users.query.get_or_404(current_user.get_id())
    user_name = user.FIO
    mols = MOLs.query.all()
    return render_template("pon_page.html", ud=ud, user_name=user_name,  mols = mols)

@app.route('/multiple_access')
@login_required
def ma_page():
    new_obj_list = get_data_for_jinja()
    user = Users.query.get_or_404(current_user.get_id())
    user_name = user.FIO
    mols = MOLs.query.all()
    return render_template("MA_page.html",  user_name=user_name, new_obj_list=new_obj_list, mols = mols )


@app.route('/buh_data')
@login_required
def buh_data_page():
    buh_data = BuhUch.query.all()
    mols = MOLs.query.all()
    user = Users.query.get_or_404(current_user.get_id())
    user_name = user.FIO
    return render_template("buh_data.html",  user_name=user_name, buh_data = buh_data, mols = mols)


@app.route('/get_data_from_db/<db>', methods=['GET', 'POST'])
@login_required
def get_data_from_db(db):
    req = request.form['json']
    print(str(datetime.now())+": запрос на получение данных -- ", db, json.loads(req))
    # db_req_lst = {
    #     'BuhUch':BuhUch.query.filter_by(inv_number=json.loads(req)).first(),
    #     # 'olt_data': get_data_for_sostav(req),
    #     'list_of_modules': List_of_modules.query.get_or_404(int(json.loads(req))),
    #     'olt_list': List_of_olt.query.get_or_404(int(json.loads(req))),
    #     'kts_data': Data_for_KTS.query.filter_by(cod_name=json.loads(req).upper()).first(),
    #     'ma_add_modules':  ma_add_modules.query.filter_by(cod_name=json.loads(req).upper()).all(),
    #     'MA_Unit_stor': MA_Units.query.filter_by(cod_name=json.loads(req).upper()).all(),
    #     'MA_Units': MA_Units.query.get_or_404(int(json.loads(req))),
    #     'Uzel_dostupa': Uzel_dostupa.query.get_or_404(int(json.loads(req))),
    #     # 'MA_Units_to_usage': get_data_for_select("MA_Units", json.loads(req)),
    #     # 'ma_add_modules_to_usage': get_data_for_select('ma_add_modules', json.loads(req)),
    #     # 'list_of_modules_move': get_data_for_move(json.loads(req))
    # }
    if db == 'BuhUch':
        buh = BuhUch.query.filter_by(inv_number=json.loads(req)).first()
        return jsonify(buh)
    if db == 'olt_data':
        return jsonify(get_data_for_sostav(req))
    if db == 'list_of_modules':
        modules = List_of_modules.query.get_or_404(int(json.loads(req)))
        return jsonify(modules)
    if db == 'olt_data_2':
        olt = List_of_olt.query.get_or_404(int(json.loads(req)))
        return jsonify(olt, olt.Type_of_olt)
    if db == 'olt_data_3':
        return jsonify(get_used_unit_data(int(json.loads(req))))
    if db == 'olt_data_4':
        
        type = Type_of_olt.query.all
        return get_data_for_new_mod()
    if db == 'kts_data':
        kts = Data_for_KTS.query.filter_by(cod_name=json.loads(req).upper()).first()
        return jsonify(kts)
    if db == 'kts_data_new':
        return get_kts_data(json.loads(req))
    if db == 'ma_add_modules':
        modules = ma_add_modules.query.filter_by(cod_name=json.loads(req).upper()).all()
        return jsonify(modules)
    if db == 'MA_Unit_stor':
        ma_units = MA_Units.query.filter_by(cod_name=json.loads(req).upper()).all()
        return jsonify(ma_units)
    if db == 'MA_Units':
        units = MA_Units.query.get_or_404(int(json.loads(req)))
        return jsonify(units, units.modules)
    if db == 'Uzel_dostupa':
        ud = Uzel_dostupa.query.get_or_404(int(json.loads(req)))
        return jsonify(ud)
    if db == 'Uzel_dostupa_all':
        ud = Uzel_dostupa.query.order_by(Uzel_dostupa.id).all()
        return ud
    if db == 'Uzel_dostupa_lst':
        ud = Uzel_dostupa.query.get_or_404(int(json.loads(req)))
        answ ={}
        for i in ud.cod_name_of_olt:
            answ[i.id] = i.cod_name_of_olt
        return jsonify(answ)
    if db == 'MA_Units_to_usage':
        # units = get_data_for_select()
        units = get_data_for_select("MA_Units", json.loads(req))
        return jsonify(units)
    if db == 'ma_add_modules_to_usage':
        units = get_data_for_select('ma_add_modules', json.loads(req))
        # units = get_data_for_select()
        return jsonify(units)
    if db == 'list_of_modules_move':
        # units = get_data_for_select()
        units = get_data_for_move(json.loads(req))
        return jsonify(units)
    if db == 'Type_of_olt':
        # units = get_data_for_select()
        data =Type_of_olt.query.all()
        return jsonify(data)
    # if db in db_req_lst:
    #     return jsonify(db_req_lst[db])
    if db == 'Objects_ur_lica':
        obj = Objects_ur_lica.query.get_or_404(int(json.loads(req)))
        units_list = {}
        modules_list = {}
        for un in range(0, len(obj.unit)) :
            units_list[un] = obj.unit[un]  
            modules_list[un]= obj.unit[un].modules    
        return jsonify(obj,units_list,modules_list)
    else:
        return None


@app.route('/delete_row/<db_name>', methods=['GET', 'POST'])
@login_required
def delete_row(db_name):
    name = request.form['json']
    id = json.loads(name)["id"]
    user = Users.query.filter_by(id=current_user.get_id()).first()
    print(str(datetime.now()) +': '+ user.FIO + " запрос на удаление -- ", db_name, name)
    db_req_lst = {
        "ma_add_modules": delMAmodules,
        "Unit":delUnit,
        "BuhUch":delBuhdata,
        "MA_Units":delMAunit,
        "Objects_ur_lica":delUrObject,
        'List_of_modules': delPONmodul,
        'Uzel_dostupa':delUD
    }
    if db_name in db_req_lst:
        return db_req_lst.get(db_name)(id)


@app.route('/change_color/<base_table>', methods=['GET', 'POST'])
@login_required
def save_color(base_table):
    name = request.form['json']
    id = json.loads(name)["id"]
    color = json.loads(name)["color"]
    print('id: '+id,', color: '+color)
    user = Users.query.filter_by(id=current_user.get_id()).first()
    if base_table == "Units":
        db_obj = Unit.query.get_or_404(int(id))
        if  db_obj.color != color:
            db_obj.color = color;
            db_obj.editor = user.FIO
        else:
            return "SUCCESS"
    if base_table == "BuhUch":
        db_obj = BuhUch.query.get_or_404(int(id))
        if  db_obj.color != color:
            db_obj.color = color;
            db_obj.editor = user.FIO
        else:
            return "SUCCESS"
    if base_table == "Objects_ur_lica":
        db_obj = Objects_ur_lica.query.get_or_404(int(id))
        if  db_obj.color != color:
            db_obj.color = color;
            db_obj.editor = user.FIO
        else:
            return "SUCCESS"    
    return save_data_to_db()


@app.route('/save_data/<db_name>', methods=['GET', 'POST'])
@login_required
def save_data(db_name):
    req_dict = json.loads(request.form['json'])
    user = Users.query.filter_by(id=current_user.get_id()).first()
    print(str(datetime.now()) +': '+ user.FIO + " запрос на внесение изменений или добавление новых записей -- ", db_name, req_dict)
    db_req_lst = {
        'KTS':save_kts_data,
        'sostav': save_sostav_data,
        'Buhuchet': save_buhuchet_data,
        'ma_add_modules_edited': save_ma_add_modules,
        'MA_Units_edited': save_ma_unit_data,
        'Unit': add_new_unit,
        'MA_Unit': add_ma_unit_data,
        'ma_add_modules': add_ma_add_modules,
        'Object_ur_lica': add_object_for_MA,
        'Objects_ur_lica_edited': save_object_for_MA,
        'list_of_modules': save_pon_modules,
        'olt_list': save_pon_olt_data,
        'olt_data_2': save_pon_olt_data,
        'Uzel_dostupa': save_ud_data,
        'add_new_pon_modules': add_new_pon_modules
    }
    if db_name in db_req_lst:
        return db_req_lst.get(db_name)(req_dict, user.FIO)


@app.route('/download/<file>/<name_PON>', methods=['GET', 'POST'])
@login_required
def downloadFile(name_PON, file):
    path = ""
    if file == "kts":
        path = create_file_KTS(name_PON)
    if file == "sostav":
        path = create_file_sostav(name_PON)
    return send_file(path, as_attachment=True)


@app.route('/download/main_table/', methods=['GET', 'POST'])
@login_required
def main_table_data_download():
    path = 'files for download/Таблица оборудования PON.xlsx'
    return send_file(path, as_attachment=True)


@app.route('/download/buh_table/', methods=['GET', 'POST'])
@login_required
def buh_table_data_download():
    path = 'files for download/Бухгалтерские данные.xlsx'
    return send_file(path, as_attachment=True)


@app.route('/buh_table_data', methods=['GET', 'POST'])
@login_required
def buh_table_data():
    name = request.form['json']
    list_data = json.loads(name)
    start_row = 4
    row_count = 1
    step = 1
    path = 'files for download\шаблон Бухгалтерские данные.xlsx'
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet = wb_obj.active
        thins = Side(border_style="thin", color="000000")
        double = Side(border_style="medium", color="000000")
        for id in list_data:
            
            buh = BuhUch.query.get_or_404(int(id))
            sheet["A" + str(start_row)] = row_count
            sheet["A" + str(start_row)].border = Border(top=double, bottom=thins, left=double, right=thins)
            sheet["B" + str(start_row)] = buh.inv_number
            sheet["B" + str(start_row)].alignment = Alignment(wrap_text=True)
            sheet["B" + str(start_row)].border = Border(top=double, bottom=thins, left=thins, right=thins)
            sheet["C" + str(start_row)] = buh.name
            sheet["C" + str(start_row)].alignment = Alignment(wrap_text=True)
            sheet["C" + str(start_row)].border = Border(top=double, bottom=thins, left=thins, right=thins)
            sheet["D" + str(start_row)] = buh.MOL
            sheet["D" + str(start_row)].border = Border(top=double, bottom=thins, left=thins, right=double)                       
            if (type(buh.charracter) == str):
                for i in range(0,len(buh.charracter.split('\n'))):
                    start = "A" + str(start_row+1+i)
                    end = "D" + str(start_row+1+i)
                    sheet.merge_cells(start+':'+end)
                    sheet["A" + str(start_row+1+i)] = buh.charracter.split('\n')[i]
                    step = len(buh.charracter.split('\n'))
            else:
                start = "A" + str(start_row+1)
                end = "D" + str(start_row+1)
                sheet.merge_cells(start+':'+end) 
                sheet["A" + str(start_row+1)] = "Нет данных!"
            start_row = start_row + 2 + step
            row_count = row_count + 1
            step = 1
        wb_obj.save('files for download\Бухгалтерские данные.xlsx')
        return json.dumps("SUCCESS")
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")
        return json.dumps(f"Unexpected {err=}, {type(err)=}")

@app.route('/main_table_data', methods=['GET', 'POST'])
@login_required
def main_table_data():
    name = request.form['json']
    list_data = json.loads(name)
    print(list_data)
    start_row = 4
    path = 'files for download\шаблон Таблица оборудования PON.xlsx'
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet = wb_obj.active
        for row in list_data:
            if(row[1] == 'List_of_olt'):
                print(row[0])
                olt = List_of_olt.query.get_or_404(str(row[0]))
                N_ud = olt.Uzel_dostupa.number_ud
                adr_ud = olt.Uzel_dostupa.Adress 
                cod = olt.cod_name_of_olt
                if  'СТЕЛАЖ' not in olt.name:
                    name = olt.name 
                else: continue
                inv_number = olt.inv_number
                ser_num = olt.serial_number
                p_mesto = ''
                note = olt.note
                r_mesto = olt.kts.mesto if olt.kts else ""
            if(row[1] == 'List_of_modules'):
                print(row[0])
                mod = List_of_modules.query.get_or_404(str(row[0]))
                N_ud = mod.List_of_olt.Uzel_dostupa.number_ud
                adr_ud = mod.List_of_olt.Uzel_dostupa.Adress
                cod = mod.List_of_olt.cod_name_of_olt
                name = mod.name_of_modules
                inv_number = mod.inv_number
                ser_num = mod.serial_number
                p_mesto = mod.Olt_sockets.socket if mod.Olt_sockets else ""
                note = mod.note
                r_mesto = mod.List_of_olt.kts.mesto if mod.List_of_olt.kts else ""
            sheet["B" + str(start_row)] = N_ud + ' ' + adr_ud
            sheet["C" + str(start_row)] = cod
            sheet["D" + str(start_row)] = name
            sheet["E" + str(start_row)] = inv_number
            sheet["F" + str(start_row)] = ser_num
            sheet["A" + str(start_row)] = r_mesto
            sheet["G" + str(start_row)] = p_mesto
            sheet["H" + str(start_row)] = note
            start_row = start_row + 1
        wb_obj.save('files for download\Таблица оборудования PON.xlsx')
        return json.dumps("SUCCESS")
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")
        return json.dumps(f"Unexpected {err=}, {type(err)=}")


def create_file_sostav(name_PON):
    unit1 = Unit.query.filter_by(name_PON=name_PON).all()
    if "UL" in name_PON:
        book = openpyxl.load_workbook('files for download\шаблон.xlsx')
        sheet = book['Huawei']
        del book['ZTE']
        plata_row = 5
    if "OL" in name_PON:
        book = openpyxl.load_workbook('files for download\шаблон.xlsx')
        sheet = book['ZTE']
        del book['Huawei']
        plata_row = 6
    for un in unit1:
        row = int(un.plata_mesto) + plata_row
        sheet['B' + str(row)] = un.inv_number
        sheet['C' + str(row)] = un.name_unit
        sheet['D' + str(row)] = un.serial_number
    book.save("files for download\Состав оборудования " + name_PON + ".xlsx")
    return "files for download\Состав оборудования " + name_PON + ".xlsx"


def create_file_KTS(name_PON):
    kts = get_kts_data(name_PON)
    # kts = Data_for_KTS.query.filter_by(cod_name=name_PON).first()
    user = Users.query.get(current_user.get_id())
    book = openpyxl.load_workbook('files for download\КТС_шаблон.xlsx')
    sheet = book.active
    sheet['A3'] = kts['full_name']
    sheet['A4'] = name_PON
    sheet['A8'] = kts['UD'] + ", " + kts['mesto'] + ", ip: " + kts['IP']
    sheet['E15'] = kts['zavod']
    sheet['L18'] = kts['date_of_production']
    sheet['G21'] = kts['Serial']
    sheet['L24'] = kts['inv_number']
    sheet['L27'] = kts['date_of_entry']
    sheet['L30'] = str(datetime.now().strftime("%d/%m/%Y"))
    sheet['J39'] = user.FIO
    book.save("files for download\КТС.xlsx")
    return "files for download\КТС.xlsx"


